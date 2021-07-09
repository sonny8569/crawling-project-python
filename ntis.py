import pandas as pd
import pyautogui
from bs4 import BeautifulSoup
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
import requests
from openpyxl import load_workbook
import openpyxl


num=1

MyId="InterYourId"#Ntis홈페이지 아이디
MyPs="InterYourPassword"#Ntis 홈페이지 비밀번호

content=[]
text=[]
driver= webdriver.Chrome()
all_values=[]
wb=openpyxl.Workbook()
sheet=wb.active
sheet.append(["회사명","수행과제","과제코드"])#과제 완료후 저장할 엑셀에 제목 미리 만들기

load_wb = load_workbook("/Users/yourPcName/Desktop/sss/test.xlsx", data_only=True)#Load The Excel File

load_ws =load_wb['Sheet1']

for row in load_ws.rows:#load_wb에서 인자를 저장합니다.
    row_value= []
    for cell in row:
        row_value.append(cell.value)
    all_values.append(row_value)


driver.get(f'https://www.ntis.go.kr/orginfo/index.jsp')#ntis로 저장
driver.find_element_by_xpath("//button[@class='login']").click()#로그인 버튼 클릭

driver.switch_to_window(driver.window_handles[1])#창화면으로 변경후 

driver.find_element_by_name('userid').send_keys(MyId)#아이디부분에 MyId 입력
driver.find_element_by_name('password').send_keys(MyPs)#비밀번호 부문에 MyPassword 입력
driver.find_element_by_xpath("//input[@type='submit']").click()
time.sleep(1)

driver.switch_to_window(driver.window_handles[0])#로그인이 끝났으니 다시 본 페이지로 돌아가기
for i in all_values:
    driver.find_element_by_id("i_searchWord").send_keys(i)#회사명 검색
    
    Key=i
    print(Key)
    pyautogui.click(593,190,button='left',clicks=1,interval=1)#중소기업 클릭 
    pyautogui.click(590,238,button='left',clicks=1,interval=1)#마우스 오토 사용

    driver.find_element_by_id("i_searchWord").send_keys(Keys.RETURN)#회사명 검색
    
    time.sleep(1)
    driver.find_element_by_xpath("//input[@type='text']").clear()#검색창 초기화
    
    try:
        driver.find_element_by_xpath("//a[@onclick='fn_srchProfilePopup(this)']").click()#회사명와 동일한 하이퍼링크 클릭
        driver.switch_to_window(driver.window_handles[1])#다시 창화면
        try:
                     
            getNewPageUrl=driver.current_url#창화면 url 가져오기

            time.sleep(1)
         
            driver.close()#기존 창화면 삭제
            driver.switch_to_window(driver.window_handles[0])#창화면 삭제후 다시 본페이지로 돌아기기
            time.sleep(1)
            CrawlingPage=webdriver.Chrome()#getNewPageUrl을 크롬으로 검색
         
            CrawlingPage.get(url=getNewPageUrl)
     
            ProjectPage=CrawlingPage.find_element_by_xpath("//li[@id='pjtList']")
            ProjectPage.click()
            pyautogui.click(520,312,button='left',clicks=2,interval=1)
            pyautogui.moveTo(1829,494)
            while True:#크롤링부분
                           
                f=open(f'result.txt','a',-1,'utf-8') 
                html=CrawlingPage.page_source
                soup=BeautifulSoup(html,'html.parser')
         
                r=soup.select('.data_box')
                k=soup.select('.resultBox')
                d=soup.select('.pjtNm')
                s=soup.select('.mdata')
            

          
                for i in k:
                        Company=Key
                        title=soup.select('.pjtNm')
                        TextFile=soup.select('.mdata')
                        f.write(Company+"*"+ title+"*"+TextFile+'\n') 
                        f.write('\n')
                        f.close()
                        wb.save("김성일.xlsx")
            
           
                   
          
            num+=1#페이지 넘기기위한 변수
            time.sleep(3)
        try:
            nextPage= CrawlingPage.find_element_by_xpath("//*[@id='pageing']/a["+str(num)+"]")#페이지 넘기기
            nextPage.click()#다음페이지 클릭
        except:#없으면 정지
            break
         

         
           
              
        
        driver.close()#다음 페이지가 없으면 종료
         
       
         

        except:
            print("수행과제 페이지가 없거나 크롤링이 끝났습니다.")
            CrawlingPage.close()
            
          

    except:
        print("페이지가 없습니다.")
driver.close()#드라이버 종료



